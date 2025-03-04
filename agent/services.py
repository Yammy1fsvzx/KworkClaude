import os
import anthropic
from django.conf import settings
import tempfile
import mimetypes
import docx
import PyPDF2
import csv
import json
from openpyxl import load_workbook

class FileProcessor:
    """
    Класс для обработки и извлечения текста из файлов различных форматов.
    
    Обеспечивает унифицированный интерфейс для извлечения текстового
    содержимого из документов разных типов для последующего анализа.
    
    Поддерживаемые форматы:
    - Текстовые файлы (.txt)
    - PDF документы (.pdf)
    - Word документы (.docx)
    - Excel таблицы (.xlsx)
    - CSV файлы (.csv)
    - JSON файлы (.json)
    """
    @staticmethod
    def extract_text_from_file(file_path):
        """
        Извлекает текстовое содержимое из файла в зависимости от его формата.
        
        Параметры:
            file_path (str): Путь к файлу для обработки
            
        Возвращает:
            str: Извлеченный текст из документа
            
        Примеры:
            >>> text = FileProcessor.extract_text_from_file('/path/to/document.pdf')
            >>> print(f"Извлечено {len(text)} символов")
        """
        mime_type, _ = mimetypes.guess_type(file_path)
        file_extension = os.path.splitext(file_path)[1].lower()
        
        # Обработка по расширению файла, если mime_type не определен
        if mime_type is None and file_extension:
            if file_extension == '.txt':
                mime_type = 'text/plain'
            elif file_extension == '.pdf':
                mime_type = 'application/pdf'
            elif file_extension == '.docx':
                mime_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            elif file_extension == '.xlsx':
                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            elif file_extension == '.csv':
                mime_type = 'text/csv'
            elif file_extension == '.json':
                mime_type = 'application/json'
        
        # Попытка прочитать как текст, если mime_type не определен или это текстовый файл
        if mime_type is None or mime_type == 'text/plain' or file_extension == '.txt':
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()
            except UnicodeDecodeError:
                # Если не удалось прочитать как UTF-8, попробуем другие кодировки
                try:
                    with open(file_path, 'r', encoding='latin-1') as f:
                        return f.read()
                except Exception as e:
                    return f"Ошибка при чтении текстового файла: {str(e)}"
                  
        elif mime_type == 'application/pdf':
            text = ""
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages:
                    text += page.extract_text() + "\n"
            return text
            
        elif mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
            doc = docx.Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
            
        elif mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            wb = load_workbook(file_path)
            text = ""
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    text += " | ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
            return text
            
        elif mime_type == 'text/csv':
            text = ""
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += " | ".join(row) + "\n"
            return text
            
        elif mime_type in ['application/json']:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return json.dumps(data, indent=2)
            
        else:
            return f"Unsupported file format: {mime_type}"

class ClaudeService:
    """
    Сервис для взаимодействия с API Claude от Anthropic.
    
    Обеспечивает функциональность для отправки запросов к Claude API,
    обрабатывает ошибки подключения и предоставляет механизм переключения
    между различными моделями Claude в случае проблем.
    
    Использование:
        ```python
        # Создание экземпляра сервиса
        claude = ClaudeService()
        
        # Сравнение документов
        result = claude.compare_documents(documents=[doc1, doc2], custom_prompt="Сравни два отчета")
        ```
    """
    def __init__(self):
        """
        Инициализация сервиса с проверкой наличия API-ключа.
        
        Загружает API-ключ из настроек проекта, инициализирует клиента
        Anthropic и настраивает список резервных моделей для автоматического
        переключения в случае ошибок.
        
        Вызывает исключение ValueError, если API-ключ не найден.
        """
        api_key = settings.CLAUDE_API_KEY
        if not api_key:
            raise ValueError("CLAUDE_API_KEY не найден в настройках. Проверьте файл .env")
        
        self.client = anthropic.Anthropic(
            api_key=api_key,
            # Увеличиваем timeout для больших запросов
            timeout=60.0
        )
        
        # Проверяем модель из настроек или используем стандартную
        self.default_model = getattr(settings, 'MODEL_NAME', 'claude-3-sonnet-20240229')
        # Альтернативные модели для автоматического переключения
        self.fallback_models = [
            'claude-3-haiku-20240307',
            'claude-3-opus-20240229',
            'claude-3-5-sonnet-20240620',
            'claude-instant-1.2',
            'claude-2.0',
        ]
        print(f"Используется модель: {self.default_model}")
    
    def _send_api_request(self, model, system_message, prompt):
        """
        Отправляет запрос к API с указанной моделью и обрабатывает ошибки.
        
        Параметры:
            model (str): Название модели Claude для использования
            system_message (str): Системное сообщение для задания контекста
            prompt (str): Основной запрос к модели
            
        Возвращает:
            tuple: (ответ от модели или None, ошибка или None)
        """
        try:
            print(f"Отправка запроса к Claude API с моделью {model}...")
            response = self.client.messages.create(
                model=model,
                max_tokens=4000,
                temperature=0,
                system=system_message,
                messages=[
                    {"role": "user", "content": prompt}
                ]
            )
            print("Ответ успешно получен!")
            return response.content[0].text, None
        except Exception as e:
            return None, e
    
    def compare_documents(self, documents, custom_prompt=None):
        """
        Анализирует список документов и отправляет их содержимое в Claude для анализа.
        
        Функция извлекает текст из каждого документа, формирует запрос к API
        Claude и возвращает результат анализа. В случае ошибок с основной моделью,
        автоматически пробует использовать резервные модели.
        
        Параметры:
            documents (QuerySet): QuerySet с объектами Document для анализа
            custom_prompt (str, optional): Пользовательский запрос для анализа
            
        Возвращает:
            str: Текстовый результат анализа от Claude
            
        Примеры:
            >>> result = claude_service.compare_documents(
            ...     documents=Document.objects.filter(id__in=['uuid1', 'uuid2']),
            ...     custom_prompt="Сравните эти документы и выделите основные различия"
            ... )
        """
        document_contents = []
        
        for doc in documents:
            print(f"Обработка документа: {doc.name} (тип: {doc.file_type})")
            with tempfile.NamedTemporaryFile(delete=False) as temp:
                temp.write(doc.file.read())
                temp_path = temp.name
            
            try:
                content = FileProcessor.extract_text_from_file(temp_path)
                file_extension = os.path.splitext(doc.name)[1].lower()
                print(f"Расширение файла: {file_extension}, определенный тип: {doc.file_type}")
                
                # Проверяем, получен ли текст
                if content and not content.startswith("Unsupported file format") and not content.startswith("Ошибка при чтении"):
                    print(f"Успешно извлечен текст из {doc.name}. Размер: {len(content)} символов")
                else:
                    print(f"Проблема с извлечением текста из {doc.name}: {content}")
                
                document_contents.append({
                    "name": doc.name,
                    "type": doc.file_type or f"Файл{file_extension}",
                    "content": content
                })
            finally:
                os.unlink(temp_path)
        
        # Prepare prompt for Claude
        if custom_prompt:
            prompt = self._build_custom_prompt(document_contents, custom_prompt)
            system_message = "You are a helpful assistant. Follow the user's instructions carefully regarding the documents."
        else:
            prompt = self._build_comparison_prompt(document_contents)
            system_message = "You are an expert analyst who performs thorough comparative analysis of documents."
        
        # Сначала пробуем основную модель
        result, error = self._send_api_request(self.default_model, system_message, prompt)
        if result:
            return result
            
        # Если основная модель не сработала, пробуем резервные модели
        print(f"Ошибка при использовании основной модели: {error}")
        print("Пробую альтернативные модели...")
        
        for fallback_model in self.fallback_models:
            print(f"Попытка использовать модель: {fallback_model}")
            result, error = self._send_api_request(fallback_model, system_message, prompt)
            if result:
                print(f"Удалось получить ответ от модели {fallback_model}")
                return result
            print(f"Модель {fallback_model} тоже не работает: {error}")
        
        # Если все модели не сработали, возвращаем сообщение об ошибке
        return f"Не удалось получить ответ ни от одной доступной модели Claude. Проверьте ваш API-ключ и доступ к моделям Claude. Последняя ошибка: {str(error)}"
    
    def _build_comparison_prompt(self, document_contents):
        """
        Формирует структурированный запрос для сравнительного анализа документов.
        
        Параметры:
            document_contents (list): Список словарей с содержимым документов
            
        Возвращает:
            str: Готовый запрос для отправки в Claude API
            
        Примечание:
            Метод ограничивает размер каждого документа до 10000 символов
            во избежание превышения лимитов API.
        """
        prompt = "Пожалуйста, проведите подробный сравнительный анализ следующих документов:\n\n"
        
        for i, doc in enumerate(document_contents, 1):
            prompt += f"## DOCUMENT {i}: {doc['name']} (Format: {doc['type']})\n\n"
            prompt += doc['content'][:10000]  # Limiting content length
            prompt += "\n\n---\n\n"
        
        prompt += """
Пожалуйста, проанализируйте эти документы и предоставьте:
1. Краткое содержание каждого документа
2. Основные сходства между документами
3. Заметные различия между документами
4. Любые идеи или закономерности, которые вы заметили
5. Выводы о том, как эти документы соотносятся друг с другом

Сформулируйте свой ответ структурированным и понятным образом, используя markdown.
"""
        return prompt
    
    def _build_custom_prompt(self, document_contents, custom_prompt):
        """
        Формирует запрос с пользовательскими инструкциями для анализа документов.
        
        Параметры:
            document_contents (list): Список словарей с содержимым документов
            custom_prompt (str): Пользовательские инструкции для анализа
            
        Возвращает:
            str: Готовый запрос для отправки в Claude API
            
        Примечание:
            Этот метод позволяет пользователям задавать собственные
            инструкции для анализа, что делает систему более гибкой.
        """
        prompt = f"I have the following documents and I need you to: {custom_prompt}\n\n"
        
        for i, doc in enumerate(document_contents, 1):
            prompt += f"## DOCUMENT {i}: {doc['name']} (Format: {doc['type']})\n\n"
            prompt += doc['content'][:10000]  # Limiting content length
            prompt += "\n\n---\n\n"
        
        prompt += """
Пожалуйста, ответьте на мой запрос, основываясь на этих документах. 
Сформулируйте свой ответ структурированным и понятным образом, используя markdown.
"""
        return prompt 